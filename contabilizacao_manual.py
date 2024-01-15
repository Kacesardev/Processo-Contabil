from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.service import Service

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains

import pymysql
import base64

import pandas as pd
import time

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment

import time
import datetime
import calendar
import os, sys

from configuracoes import *
import gravar_log_database

start = time.time()
status_script = "Ok"

# Defininição de variáveis
descricao_log = "Contabilização Manual"
codigo_script_log = 9 # Variavel de identificacao do Script no Banco de Dados - Tabela Catalogo
error_response = ''
DEBUG = 1
driver = None



regionais_lista = {'itaipu':'Itaipu','itaipunorte':'Itaipu Norte','equipo':'Equipo','quintaroda':'Quinta Roda', 'wlm-sede':'WLM-Sede', 'csc':'CSC'}
opcoes_empresa = {'itaipu':'WLM - REGIONAL MINAS','itaipunorte':'WLM - REGIONAL NORTE','equipo':'WLM - REGIONAL RIO','quintaroda':'WLM -  REGIONAL SAO PAULO', 'wlm-sede':'WLM - MATRIZ', 'csc':'WLM - CSC-MG'}
arquivo = "Robô Contabilização Manual Oficial.xlsx"
local_arquivo = r"\\" + pasta_rede + "\itaipu-fs\Interdep\CSC-Contabilidade\Contabilidade - CSC\Contabilização Automática\\"
url_login = "https://siconnet.scania.com.br/sicomnet3/wlm/sicomweb.gen.gen.pag.Login.cls"

HOST_NAME = '172.10.10.8'
USER_NAME = 'aW1fYWNlc3Nv'
PASWD = 'U2NhbmlhQDIwMTk='

# Verificação dos parâmetros iniciais para inicialização do script
if len(sys.argv) >= 2:
    var_regional = sys.argv[1]
else:
    print("Parâmetros inválidos acesse com: Regional")
    sys.exit()

def decod(var):
    # Decodificar base64
    text_var = base64.b64decode(var.encode('ascii'))
    return text_var.decode('ascii')

def carrega_credencial(tipo_acesso):
    # Carregar usuário e senha do SiconNet utilizado pelo script: WLMAUT
    try:
        con = pymysql.connect(host=HOST_NAME, user=decod(USER_NAME), password=decod(PASWD), db='automacao_processos', charset='utf8')
        query_db = con.cursor()
    except Exception as e:
        print("Erro: Impossível conectar ao Banco de Dados, Erro: " + str(e))

    try:
        sql_command = "SELECT login, senha FROM cofre_acessos WHERE descricao=%s"
        query_db.execute(sql_command, (tipo_acesso,))
        row_db = query_db.fetchone()
    except Exception as e:
        print("Erro: Impossível acessar a tabela de acessos, Erro: " + str(e))
    finally:
        query_db.close()
        con.close()
    return row_db[0], row_db[1]

# Carrega as credenciais do arquivo de configurações
user_sicon, pwd_sicon = carrega_credencial('SiconNet')

def send_email(subject_status, message):
    # Enviar Email
    smtp_port = 587
    sender_email = 'sistemas@wlm.com.br'
    receiver_email = 'contabilidade.csc@wlm.com.br'
   

    # Criar o objeto de email
    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = 'Robô Contábil - Contabilização Manual - Regional ' + regionais_lista[var_regional] + ' : ' + str(subject_status)

    # Adicionar o corpo da mensagem
    body = message
    msg.attach(MIMEText(body, 'plain'))

    # Enviar o email
    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(smtp_user, smtp_pass)
            server.send_message(msg)
        print("Email enviado com sucesso") if DEBUG == 1 else ""
    except Exception as e:
        print(f"Erro ao enviar o email: {str(e)}")

def gravar_excel(valor_celula, posicao_linha):
    # Atualizar a coluna do Excel confirmando ou não o lançamento contábil
    wb = load_workbook(local_arquivo + regionais_lista[var_regional] + "\\" + arquivo)
    ws = wb['Lançamentos Contábeis']
    ws['T' + str(posicao_linha)].value = str(valor_celula)
    try:
        wb.save(local_arquivo + regionais_lista[var_regional] + "\\" + arquivo)
        print('Excel atualizado') if DEBUG == 1 else ""
    except:
        print("Erro ao salvar no arquivo Excel, verifique se o arquivo esta aberto.")

def abre_browser():
    # Abrir o navegador
    print("browser sendo aberto.") if DEBUG == 1 else ""
    options = Options()
    options.add_argument("--start-maximized")
    options.add_argument("--window-size=1920,1080")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--disable-automation")
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    options.add_argument("--disable-gpu")
    options.add_argument("--disable-extensions")
    options.add_argument("--proxy-server='direct://'")
    options.add_argument("--proxy-bypass-list=*")
    options.add_argument("--start-maximized")
    options.add_argument("--disable-notifications")
    options.add_argument("--disable-popup-blocking")
    options.add_argument("--disable-notifications")  # Desabilitar notificações
    driver = webdriver.Chrome(service=Service(), options=options)
    return driver

def entra_login(driver, url_login):
    # Entrar no login
    print("login será realizado.") if DEBUG == 1 else ""
    driver.get(url_login)
    return driver

def aceita_alerta(driver):
    # Aceitar alerta
    try:
        print("aceitando alerta.") if DEBUG == 1 else ""
        WebDriverWait(driver, 5).until(EC.alert_is_present(),'Timed out waiting for PA creation ')
        alert = driver.switch_to.alert
        alert.accept()
    except:
        pass
    return driver

def faz_login(driver, login, senha):
    # Fazer login
    print("fazendo login.") if DEBUG == 1 else ""
    input_usuario = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "control_11")))
    input_senha = WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.ID, "control_15")))
    input_usuario.click()
    input_usuario.clear()
    input_usuario.send_keys(login)
    input_senha.click()
    input_senha.clear()
    input_senha.send_keys(senha)
    botao = driver.find_element(By.ID, 'botaoLogin')
    botao.click()
    return driver

def faz_logoff(driver):
    # Fazer logoff
    print("fazendo logoff.") if DEBUG == 1 else ""
    time.sleep(3)
    driver.switch_to.default_content()
    # Clicar no botão logoff
    driver.find_element(By.ID, "image_33").click()
    driver.switch_to.window(driver.window_handles[0])
    elemento = WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.ID, "zen29")))
    driver.close()
    return driver

def entra_lac_contab(driver):
    # Entrar na rotina do SiconNet: lançamento contábil
    print("entrando em lançamento contabil.") if DEBUG == 1 else ""
    driver.switch_to.window(driver.window_handles[1])
    javascript_code = "zenPage.carregaFrame('sicomweb.ctm.mv.pag.LancamentoContabLote.cls', 552, 0);"
    driver.execute_script(javascript_code)
    return driver

def seleciona_dropdown(driver):
    # Selecionar o dropdown
    print("selecionando dropdown.") if DEBUG == 1 else ""
    iframe = driver.find_element(By.ID, 'iframe_34')
    driver.switch_to.frame(iframe)
    driver.find_element(By.ID, "btn_13").click()
    driver.execute_script("zenPage.getComponent(13).showDropdown();")
    return driver

def captura_opcoes(driver):
    # Capturar opções
    print("capturando opções de empresa.") if DEBUG == 1 else ""
    combo = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.CSS_SELECTOR, 'table[class="comboboxTable"]')))
    combo = combo.text
    lista_combo = []
    combo_split = combo.split("\n")
    for i, texto in enumerate(combo_split):
        if i % 2 == 0:
            lista_combo.append(texto)
    print("Selecione a empresa desejada:") if DEBUG == 1 else ""
    selecao_empresa = input(lista_combo).replace("'", "")
    return selecao_empresa

def seleciona_empresa(driver, selecao_empresa):
    print(f"selecionando empresa. {selecao_empresa}") if DEBUG == 1 else ""
    time.sleep(5)
    # Seleciona a empresa
    driver.find_element(By.CSS_SELECTOR, f"tr[zentext='{selecao_empresa}']").click()
    driver.find_element(By.CSS_SELECTOR, 'body[id="zenBody"]').click()
    # Botao OK
    botao_ok = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, 'control_14'))).click()
    return driver

def efetua_lancamento(driver):
    # Obter os dados do arquivo
    print("obtendo dados do arquivo.") if DEBUG == 1 else ""
    try:
        df = pd.read_excel(local_arquivo + regionais_lista[var_regional] + "\\" + arquivo, dtype="str")
    except:
        print("Não foi possível encontrar o arquivo " + nome_arquivo + ", verifique.")
        sys.exit()

    # Ajusta dados lidos da planilha para processamento
    df["Valor"] = df["Valor"].astype(float)
    df["Valor Total"] = df["Valor Total"].astype(float)
    df["Valor"] = df["Valor"].apply(lambda x: '{:.2f}'.format(x))
    df["Valor Total"] = df["Valor Total"].apply(lambda x: '{:.2f}'.format(x))
    df.fillna('', inplace=True)
    df.dropna(subset=['Lote'], inplace=True) # Remover linhas com lote vazio

    qtd_linhas = len(df.index)
    qtd_lotes =  len(df.groupby('Lote').nunique())
    print(f"Resumo: \n Total de Lotes {qtd_lotes}, Total de Itens: {qtd_linhas} \n")

    # Variaveis auxiliares para incluir no corpo do email para usuário.
    qtd_linhas_erro = 0
    qtd_linhas_lancadas = 0
    lotes_processados = []

    # Listar os Lotes para processar
    for lote_index, df_itens in df.groupby(['Lote'], sort=False):
        print(f'\nLote: {lote_index[0]} ')
        lotes_processados.append(str(lote_index[0]))
        
        # Primeira Tela -> registrar: lote, numero de lançamentos, valor total, dia
        print("1a parte: dados do lote.") if DEBUG == 1 else ""
        
        # Lote
        input_lote = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, 'control_60')))
        input_lote.click()
        input_lote.clear()
        input_lote.send_keys(lote_index[0])
        input_lote.send_keys(Keys.TAB)
        
        botao_gravar = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, 'control_16'))).click()
        
        # Confirmar após alerta
        try:
            print("aceitando alerta.") if DEBUG == 1 else ""
            WebDriverWait(driver, 3).until(EC.alert_is_present(),'Timed out waiting for PA creation ')
            alert = driver.switch_to.alert
            alerta_mensagem = alert.text
            if alerta_mensagem == "Dados atualizados com sucesso.":
                # confirmar caso o sistema responda que está ok.
                alert.accept()
            else:
                # erro ir para o próximo lote
                print("Erro, não foi possível gravar o lote com sucesso.")
                break
        except:
            pass
        
        # Clicar no link: Lançamentos -> navegar para a segunda tela
        print("abrindo tela lançamentos.") if DEBUG == 1 else ""
        driver.find_element(By.CSS_SELECTOR, 'a.link#a_27').click()
        
        for subitens_index, df_sub in df_itens.iterrows():
            # Segunda Tela -> registrar lançamentos do lote: conta débito ou conta crédito, valor, historico digitado, centro de custo (Se houver), conta corrente
            print("2a parte: lançamentos contábeis.") if DEBUG == 1 else ""

            posicao_excel = int(subitens_index) + 2 # Marca a posição da linha do Excel considerando o titulo

            if df_sub['Status'] != "Ok":
                # Código da Conta de Débito
                input_conta_debito = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, 'control_60')))
                input_conta_debito.click()
                input_conta_debito.clear()
                input_conta_debito.send_keys(df_sub['Conta Débito'])

                # Código da Conta de Crédito
                input_conta_credito = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, 'control_61')))
                input_conta_credito.click()
                input_conta_credito.clear()
                input_conta_credito.send_keys(df_sub['Conta Crédito'])

                # Valor
                input_valor = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, 'control_62')))
                input_valor.click()
                input_valor.clear()
                input_valor.send_keys(df_sub['Valor'])

                # Option: Tipo do Histórico = Digitado
                driver.find_element(By.ID, 'caption_1_53').click()
                
                # Descrição do Histórico
                input_digitado = driver.find_element(By.ID, 'control_64')
                input_digitado.click()
                input_digitado.clear()
                input_digitado.send_keys(df_sub['Digitado'])
                
                # Centro de Custos: Código Reduzido
                try:
                    input_centro_custo = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.ID, 'codigoRed_66')))
                    input_centro_custo.click()
                    input_centro_custo.clear()
                    input_centro_custo.send_keys(df_sub['Centro de Custos'])
                except:
                    pass
                
                # Centro de Custos: Conta Corrente
                try:
                    input_conta_corrente = WebDriverWait(driver, 3).until(EC.element_to_be_clickable((By.ID, 'control_66')))
                    input_conta_corrente.click()
                    input_conta_corrente.send_keys(Keys.TAB)
                except:
                     pass
                
                # Gravar
                botao_gravar = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, 'control_16'))).click()

                # Confirmar após alerta: Caso apareça algum alerta aqui é porque alguma informação ficou faltando ou está errada, gravar essa mensagem no Excel e continuar
                try:
                    print("aceitando alerta.") if DEBUG == 1 else ""
                    WebDriverWait(driver, 3).until(EC.alert_is_present(),'Timed out waiting for PA creation ')
                    alert = driver.switch_to.alert
                    alerta_mensagem = alert.text
                    alert.accept()
                    gravar_excel('Erro: ' + str(alerta_mensagem) , posicao_excel)
                    qtd_linhas_erro += 1
                    # erro ir para o próximo lote
                    print("Erro, não foi possível gravar o lançamento.")
                    continue
                except:
                    # Lançamento Ok -> Atualizar e coluna Status do Excel na célula corresponte ao lançamento atual
                    gravar_excel('Ok', posicao_excel)
                    qtd_linhas_lancadas += 1

                # Botão Novo
                botao_novo = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((By.ID, 'control_18'))).click()

            else:
                print("Item já lançado anteriomente, nada realizado.")
                    
        print("retornando a primera tela.") if DEBUG == 1 else ""
        driver.find_element(By.ID,'a_28').click() # Botão Retornar
    texto_retorno = 'Qtd de lançamentos com sucesso: ' + str(qtd_linhas_lancadas) + ', com erro: ' + str(qtd_linhas_erro) + '\n\nLotes processados: ' + ', '.join(lotes_processados)
    return texto_retorno

def iniciar_script():
    # Função Principal que executa as etapas do processo.
    global status_script, driver, df, error_response
    try:
        # Abrir o navegador
        driver = abre_browser()

        # Fazer Login no SiconNet
        driver = entra_login(driver, url_login)
        driver = aceita_alerta(driver) # no ambiente de aceite é necessário aceitar alerta, timeout 5 segundos
        driver = faz_login(driver, user_sicon, pwd_sicon)

        # Entrar na rotina
        time.sleep(5)
        driver = entra_lac_contab(driver)

        # Selecionar a empresa
        time.sleep(5)
        driver = seleciona_dropdown(driver)
        time.sleep(5)
        driver = seleciona_empresa(driver, opcoes_empresa[var_regional])

        # Iniciar Processo de Lançamentos
        texto_retorno = efetua_lancamento(driver)

        # Enviar Email sobre a conclusão do processo
        mensagem_email = "Script de Contabilização Manual concluído,\n" + texto_retorno
        send_email("Concluído", mensagem_email)

        # Fazer Logoff do SiconNet e fechar tela
        driver = faz_logoff(driver)
        
    except Exception as e:
        error_response = f"Informações sobre o erro:\n{e}"
        status_script = "Not"
        send_email("Erro", error_response)
    finally:
        try:
            driver.close()
        except:
            pass
        driver.quit()

# Executar código no loop principal
print(descricao_log)

###send_email("Inicio","Script de Contabilização Manual foi iniciada.")
iniciar_script()


end = time.time()
time_spended = round(end - start,2)
print("Finalizado! ->" + str(time_spended) + "s")
gravar_log_database.gravar_log_database(codigo_script_log, time_spended, status_script, sys.argv, error_response)
